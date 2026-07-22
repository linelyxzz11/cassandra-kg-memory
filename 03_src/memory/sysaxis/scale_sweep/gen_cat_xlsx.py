import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path("D:/memorytable/cassandra-kg-memory/reports/locomo_cat1_4_detailed.xlsx")
wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def style_data(ws, start_row, end_row, cols, best_cols=None):
    for r in range(start_row, end_row+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = thin_border

# ======== Sheet 1: Retrieval ========
ws1 = wb.active
ws1.title = "Retrieval cat1-4"

cats = [("cat1 multi-hop", 282), ("cat2 temporal", 321), ("cat3 commonsense", 96), ("cat4 single-hop", 841)]

# Per-category sheets
for ci, (cname, cn) in enumerate(cats):
    cat_label = f"cat{ci+1}"
    ws1.cell(row=1, column=ci*7+1, value=f"=== {cname} (n={cn}) ===")
    ws1.merge_cells(start_row=1, start_column=ci*7+1, end_row=1, end_column=ci*7+6)
    ws1.cell(row=1, column=ci*7+1).font = Font(bold=True, size=12)
    
    headers = ["Method", "R@1", "R@5", "R@10", "MRR", "vs BM25 ΔR@1"]
    for j, h in enumerate(headers):
        ws1.cell(row=2, column=ci*7+1+j, value=h)
    style_header(ws1, 2, ci*7+6)

    data = [
        ("BM25", 0.1028 if ci==0 else 0.3209 if ci==1 else 0.0833 if ci==2 else 0.3008,
         0.2943 if ci==0 else 0.5358 if ci==1 else 0.2708 if ci==2 else 0.5184,
         0.4184 if ci==0 else 0.6137 if ci==1 else 0.3125 if ci==2 else 0.5945,
         0.1899 if ci==0 else 0.4202 if ci==1 else 0.1652 if ci==2 else 0.3958),
        ("Dense-bge", 0.2801 if ci==0 else 0.4766 if ci==1 else 0.1875 if ci==2 else 0.3876,
         0.6277 if ci==0 else 0.7040 if ci==1 else 0.3542 if ci==2 else 0.6564,
         0.7447 if ci==0 else 0.7539 if ci==1 else 0.4479 if ci==2 else 0.7479,
         0.4220 if ci==0 else 0.5704 if ci==1 else 0.2504 if ci==2 else 0.5020),
        ("Dense+GlobalKG", 0.3617 if ci==0 else 0.4953 if ci==1 else 0.2604 if ci==2 else 0.4174,
         0.6809 if ci==0 else 0.7072 if ci==1 else 0.3854 if ci==2 else 0.6433,
         0.7801 if ci==0 else 0.7601 if ci==1 else 0.5208 if ci==2 else 0.7325,
         0.4885 if ci==0 else 0.5878 if ci==1 else 0.3141 if ci==2 else 0.5142),
        ("Dense+QueryKG", 0.2908 if ci==0 else 0.5140 if ci==1 else 0.2083 if ci==2 else 0.4043,
         0.6348 if ci==0 else 0.7290 if ci==1 else 0.3854 if ci==2 else 0.6790,
         0.7660 if ci==0 else 0.7632 if ci==1 else 0.4792 if ci==2 else 0.7693,
         0.4346 if ci==0 else 0.6074 if ci==1 else 0.2777 if ci==2 else 0.5211),
    ]
    
    bm25_r1 = data[0][1]
    for i, (m, r1, r5, r10, mrr) in enumerate(data):
        ws1.cell(row=3+i, column=ci*7+1, value=m)
        for j, v in enumerate([r1, r5, r10, mrr]):
            ws1.cell(row=3+i, column=ci*7+2+j, value=v)
            ws1.cell(row=3+i, column=ci*7+2+j).number_format = '0.0000'
        delta = r1 - bm25_r1 if i > 0 else 0
        ws1.cell(row=3+i, column=ci*7+6, value=f"{delta:+.4f}" if i > 0 else "—")
    style_data(ws1, 3, 6, ci*7+6)
    
    # Highlight best in each column
    best_r1 = max(data[1][1] for d in data[1:])  # max R@1 excluding BM25
    for i in range(1, 4):
        if data[i][1] == best_r1:
            ws1.cell(row=3+i, column=ci*7+2).fill = best_fill

# ======== Sheet 2: Retrieval Overview ========
ws2 = wb.create_sheet("Retrieval Overview")
ws2.cell(row=1, column=1, value="Method")
ws2.cell(row=1, column=2, value="Overall R@1")
ws2.cell(row=1, column=3, value="Overall R@10")
ws2.cell(row=1, column=4, value="Overall MRR")
for ci, (cn, _) in enumerate(cats):
    ws2.cell(row=1, column=5+ci*3, value=f"{cn} R@1")
    ws2.cell(row=1, column=6+ci*3, value=f"{cn} R@10")
    ws2.cell(row=1, column=7+ci*3, value=f"{cn} MRR")
style_header(ws2, 1, 16)

overall_data = [
    ("BM25", 0.2649, 0.5619, 0.3600,
     [(0.1028, 0.4184, 0.1899), (0.3209, 0.6137, 0.4202), (0.0833, 0.3125, 0.1652), (0.3008, 0.5945, 0.3958)]),
    ("Dense-bge", 0.3419, 0.7009, 0.4534,
     [(0.2801, 0.7447, 0.4220), (0.4766, 0.7539, 0.5704), (0.1875, 0.4479, 0.2504), (0.3876, 0.7479, 0.5020)]),
    ("Dense+GlobalKG", 0.3872, 0.7095, 0.4851,
     [(0.3617, 0.7801, 0.4885), (0.4953, 0.7601, 0.5878), (0.2604, 0.5208, 0.3141), (0.4174, 0.7325, 0.5142)]),
    ("Dense+QueryKG", 0.3585, 0.7185, 0.4724,
     [(0.2908, 0.7660, 0.4346), (0.5140, 0.7632, 0.6074), (0.2083, 0.4792, 0.2777), (0.4043, 0.7693, 0.5211)]),
]

for i, (m, r1, r10, mrr, cat_data) in enumerate(overall_data):
    ws2.cell(row=2+i, column=1, value=m)
    ws2.cell(row=2+i, column=2, value=r1)
    ws2.cell(row=2+i, column=3, value=r10)
    ws2.cell(row=2+i, column=4, value=mrr)
    for ci, (cr1, cr10, cmrr) in enumerate(cat_data):
        ws2.cell(row=2+i, column=5+ci*3, value=cr1)
        ws2.cell(row=2+i, column=6+ci*3, value=cr10)
        ws2.cell(row=2+i, column=7+ci*3, value=cmrr)
style_data(ws2, 2, 5, 16)

# Highlight best in overall columns
for col in [2, 3, 4]:
    values = [(i, ws2.cell(row=2+i, column=col).value) for i in range(4)]
    best_val = max(v for _, v in values)
    for i, v in values:
        if abs(v - best_val) < 0.001:
            ws2.cell(row=2+i, column=col).fill = best_fill

# ======== Sheet 3: Reader cat1-4 ========
ws3 = wb.create_sheet("Reader cat1-4")
ws3.cell(row=1, column=1, value="Method")
ws3.cell(row=1, column=2, value="Overall rF1")
ws3.cell(row=1, column=3, value="Overall rEM")
ws3.cell(row=1, column=4, value="Overall WrongAbst")
ws3.cell(row=1, column=5, value="Overall Hit@10")
for ci, (cn, cn_n) in enumerate(cats):
    ws3.cell(row=1, column=6+ci*3, value=f"{cn} rF1")
    ws3.cell(row=1, column=7+ci*3, value=f"{cn} Hit@10")
    ws3.cell(row=1, column=8+ci*3, value=f"{cn} WrongAbst")
style_header(ws3, 1, 17)

reader_data = [
    ("BM25", 0.2648, 0.1539, 0.5929, 0.5487,
     [(0.1367, 0.4184, 0.7411), (0.1396, 0.6137, 0.6480), (0.0122, 0.3125, 0.9583), (0.3845, 0.5945, 0.4804)]),
    ("Dense-bge", 0.3482, 0.1831, 0.4539, 0.7299,
     [(0.2698, 0.7447, 0.4752), (0.1553, 0.7539, 0.6262), (0.0481, 0.4479, 0.8958), (0.4824, 0.7479, 0.3306)]),
    ("Dense+GlobalKG", 0.3222, 0.1701, 0.4877, 0.7338,
     [(0.2569, 0.7801, 0.4787), (0.1560, 0.7601, 0.6417), (0.0523, 0.5208, 0.8750), (0.4384, 0.7325, 0.3876)]),
    ("Dense+QueryKG", 0.3516, 0.1877, 0.4494, 0.7494,
     [(0.2654, 0.7660, 0.4716), (0.1658, 0.7632, 0.6012), (0.0640, 0.4792, 0.8646), (0.4844, 0.7693, 0.3365)]),
]

for i, (m, rf1, rem, wa, hit, cat_data) in enumerate(reader_data):
    ws3.cell(row=2+i, column=1, value=m)
    ws3.cell(row=2+i, column=2, value=rf1)
    ws3.cell(row=2+i, column=3, value=rem)
    ws3.cell(row=2+i, column=4, value=wa)
    ws3.cell(row=2+i, column=5, value=hit)
    for ci, (crf1, chit, cwa) in enumerate(cat_data):
        ws3.cell(row=2+i, column=6+ci*3, value=crf1)
        ws3.cell(row=2+i, column=7+ci*3, value=chit)
        ws3.cell(row=2+i, column=8+ci*3, value=cwa)
style_data(ws3, 2, 5, 17)

for col in [2, 5]:
    values = [(i, ws3.cell(row=2+i, column=col).value) for i in range(4)]
    best_val = max(v for _, v in values)
    for i, v in values:
        if abs(v - best_val) < 0.001:
            ws3.cell(row=2+i, column=col).fill = best_fill

# Column widths
for ws, widths in [(ws1, [16,10,10,10,10,14]*4), (ws2, [18,12,12,12]+[12,12,12]*4), (ws3, [18,12,10,14,12]+[12,12,14]*4)]:
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

wb.save(OUT)
print(f"Written: {OUT} ({OUT.stat().st_size} bytes)")
